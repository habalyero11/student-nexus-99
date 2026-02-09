import openpyxl
from openpyxl import load_workbook
import json

# Load the Excel file
wb = load_workbook('7-LAPLACE-ESP.xlsx')
sheet = wb.active

# Create analysis dictionary
analysis = {
    "sheet_name": sheet.title,
    "dimensions": sheet.dimensions,
    "max_row": sheet.max_row,
    "max_column": sheet.max_column,
    "metadata_rows": [],
    "column_headers": {},
    "sample_data": [],
    "formulas": [],
    "merged_cells": []
}

# Get first 20 rows for metadata and headers
print("Reading header rows...")
for row_idx in range(1, min(21, sheet.max_row + 1)):
    row_data = {}
    for col_idx in range(1, min(sheet.max_column + 1, 30)):
        cell = sheet.cell(row=row_idx, column=col_idx)
        if cell.value:
            row_data[f"col_{col_idx}"] = str(cell.value)[:200]  # Limit length
    if row_data:
        analysis["metadata_rows"].append({
            "row": row_idx,
            "data": row_data
        })

# Find column headers (look for specific keywords)
print("Finding column headers...")
header_row_idx = None
for row_idx in range(1, min(25, sheet.max_row + 1)):
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = str(sheet.cell(row=row_idx, column=col_idx).value or "").upper()
        if any(keyword in cell_value for keyword in ["WRITTEN WORK", "PERFORMANCE TASK", "QUARTERLY"]):
            header_row_idx = row_idx
            print(f"Found header row at: {row_idx}")
            break
    if header_row_idx:
        break

if header_row_idx:
    for col_idx in range(1, min(sheet.max_column + 1, 50)):
        cell = sheet.cell(row=header_row_idx, column=col_idx)
        if cell.value:
            analysis["column_headers"][f"col_{col_idx}"] = str(cell.value)[:200]

# Get sample data (5 rows after header)
print("Reading sample data...")
if header_row_idx:
    for row_idx in range(header_row_idx + 1, min(header_row_idx + 6, sheet.max_row + 1)):
        row_data = {}
        for col_idx in range(1, min(15, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                row_data[f"col_{col_idx}"] = str(cell.value)[:100]
        if row_data:
            analysis["sample_data"].append({
                "row": row_idx,
                "data": row_data
            })

# Find formulas
print("Finding formulas...")
for row_idx in range(1, min(sheet.max_row + 1, 30)):
    for col_idx in range(1, min(sheet.max_column + 1, 30)):
        cell = sheet.cell(row=row_idx, column=col_idx)
        if cell.data_type == 'f':
            analysis["formulas"].append({
                "cell": cell.coordinate,
                "formula": str(cell.value)[:300]
            })
            if len(analysis["formulas"]) >= 30:
                break
    if len(analysis["formulas"]) >= 30:
        break

# Get merged cells
print("Finding merged cells...")
if sheet.merged_cells:
    for merged_range in list(sheet.merged_cells.ranges)[:30]:
        analysis["merged_cells"].append(str(merged_range))

# Write to JSON file
print("Writing analysis to JSON...")
with open('excel_structure.json', 'w', encoding='utf-8') as f:
    json.dump(analysis, f, indent=2, ensure_ascii=False)

print("\nAnalysis complete! Check excel_structure.json")
print(f"Total rows: {analysis['max_row']}")
print(f"Total columns: {analysis['max_column']}")
print(f"Header row: {header_row_idx}")
print(f"Formulas found: {len(analysis['formulas'])}")

wb.close()
